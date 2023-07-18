#!/usr/bin/env python
# coding: utf-8

# In[19]:


####################### Quote to csv converter 25APR23 ########################
# Just dump all the official excel files from ***** and get them collected  #
# in a command separated value file                                           #
###############################################################################
import pandas as pd
import os
from datetime import datetime
import openpyxl
import re
import csv
date_format = "%d-%b-%Y"


# In[20]:


get_ipython().system('rm inventory.csv')


# In[21]:


files=os.listdir('.')


# In[22]:


f = open('inventory.csv','w')
line='file;Account;Quote;End_user;Support _Coverage;Product_name;Start_date;End_date;SSRN;delta.days;Price List Calculated;Price List quoted;Discount;Cost'
f.write(line+"\n")


# In[23]:


Row=20
Support_Coverage="1"


for file in files:
    if "xlsx" in file:
        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            while Support_Coverage != "":
                Support_Coverage =ws.cell(row = Row, column = 1).value
                Account          =ws.cell(row = 10,  column = 2).value
                Quote            =ws.cell(row = 2 ,  column = 14).value
                End_user         =ws.cell(row = 10,  column = 6).value
                Product_name     =ws.cell(row = Row, column = 2).value
                QTY              =ws.cell(row = Row, column = 3).value
                Start_date       =ws.cell(row = Row, column = 4).value
                End_date         =ws.cell(row = Row, column = 5).value
                delta            = datetime.strptime(End_date, date_format)-datetime.strptime(Start_date, date_format)
                h                =(363/(delta.days)) #year conversion
                for i in range(int(QTY)):
                    SSRN      =ws.cell(row = Row+i+1, column = 6).value                    
                    Installed =ws.cell(row = Row+i+1, column = 7).value
                    PL        =ws.cell(row = Row, column = 13).value
                    PL        =float(PL.replace(',', ''))                    #fixing the comma
                    Discount  =ws.cell(row = Row, column = 16).value
                    Discount  =float(Discount.replace('%', ''))/100          #fixing the comma and making a unit
                    line      =file+";"+Account +";"+str(Quote)+";"+End_user+";"+Support_Coverage+";"+Product_name+";"+ Start_date+";"\
                    +End_date+";'"+SSRN+";"+str(delta.days)+";"+str(PL*h)+";"+str(PL)+";"+str(Discount)+";"+str(PL*(1+Discount))
                    print((line))
                    f.write(line+"\n")
                Row+=int(QTY)+1
                Support_Coverage=ws.cell(row = Row, column = 1).value
        except:   #this will break when finding a nan as number for the next row
            Row=20
            pass
                
                


# In[24]:


f.close()


# In[ ]:





# In[ ]:


get_ipython().system('mkdir ../old')


# In[ ]:




